# -*- coding: utf-8 -*-
import logging

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger()


class exceloutput:
    def __init__(self, excel_path, tabs, updatetabs):
        self.workbook = Workbook()
        self.excel_path = excel_path
        self.tabs = tabs
        self.updatetabs = updatetabs

    def update_tab(self, tabname, values):
        if tabname not in self.updatetabs:
            return
        sheetname = self.tabs[tabname]
        try:
            del self.workbook[sheetname]
        except KeyError:
            pass
        tab = self.workbook.create_sheet(sheetname)
        if isinstance(values, list):
            for i, row in enumerate(values):
                for j, value in enumerate(row):
                    tab.cell(row=i+1, column=j+1, value=value)
        else:
            for r in dataframe_to_rows(values, index=True, header=True):
                tab.append(r)

    def save(self):
        self.workbook.save(self.excel_path)
